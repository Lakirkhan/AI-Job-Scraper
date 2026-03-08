"""
Excel Exporter Module
Exports job listings to Excel format with formatting
"""

import pandas as pd
from datetime import datetime
from utils.config import OUTPUT_FOLDER, EXCEL_PREFIX
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

class ExcelExporter:
    def __init__(self):
        self.output_folder = OUTPUT_FOLDER
        self._ensure_output_folder()
    
    def _ensure_output_folder(self):
        """Create output folder if it doesn't exist"""
        if not os.path.exists(self.output_folder):
            os.makedirs(self.output_folder)
            print(f"✓ Created output folder: {self.output_folder}")
    
    def export(self, jobs, filename=None):
        """
        Export jobs to formatted Excel file
        
        Args:
            jobs (list): List of job dictionaries
            filename (str): Custom filename (optional)
            
        Returns:
            str: Path to saved file
        """
        if not jobs:
            print("⚠ No jobs to export to Excel")
            return None
        
        try:
            # Create filename
            if not filename:
                timestamp = datetime.now().strftime('%Y%m%d_%H%M')
                filename = f"{EXCEL_PREFIX}{timestamp}.xlsx"
            
            filepath = os.path.join(self.output_folder, filename)
            
            # Convert to DataFrame
            df = pd.DataFrame(jobs)
            
            # Create Excel file
            df.to_excel(filepath, index=False, sheet_name='Jobs')
            
            # Format the Excel file
            self._format_excel(filepath)
            
            print(f"✓ Excel exported: {filepath} ({len(jobs)} jobs)")
            return filepath
            
        except Exception as e:
            print(f"✗ Error exporting to Excel: {str(e)}")
            return None
    
    def _format_excel(self, filepath):
        """Apply formatting to Excel file"""
        try:
            workbook = load_workbook(filepath)
            worksheet = workbook.active
            
            # Header formatting
            header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # Border style
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Format header row
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # Format data rows
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value or '')) > max_length:
                            max_length = len(str(cell.value or ''))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Set row height for header
            worksheet.row_dimensions[1].height = 25
            
            workbook.save(filepath)
            
        except Exception as e:
            print(f"⚠ Warning: Could not apply formatting: {str(e)}")

def export_to_excel(jobs, filename=None):
    """Backward compatible function"""
    exporter = ExcelExporter()
    return exporter.export(jobs, filename)